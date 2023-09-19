#!/usr/bin/env python
from flask import Flask, flash, session, url_for, render_template, request, redirect, jsonify
from flask_session import Session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash # Password Hash
from flask_bcrypt import Bcrypt
from app import User, Visitor, Card, User_log, Year, Month, Day, Department, Rack, Privacy, Password_log, Account_log, Login_failure_log, Permission_log, Privacy_log, Password_change_log
from pycrypto import *
import jinja2.exceptions
from config import create_app, db
from flask_login import login_user, login_required, logout_user, current_user
from flask_login import LoginManager
import datetime
from datetime import datetime, date, time, timedelta
from sqlalchemy import func, Integer, and_, or_
import qrcode
import mysql.connector
import openpyxl
import json
import pandas as pd
from tabulate import tabulate
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import re
from flask.helpers import get_flashed_messages
from db_connector import db_connector, account_list
import pytz
import string
import random
import hashlib
import smtplib
from email.mime.text import MIMEText
from apscheduler.schedulers.background import BackgroundScheduler
import os

# 암복호화 로직
key = "avhejrghjawerjvawev"
aes = AESCipher(key)

app = create_app()
bcrypt = Bcrypt(app)

SECRET_TOKEN = 'Ssw8Ik0OWZM3vYSyqPQdgo1M7oPSE5EZLayiMPPlMFg'

# Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = u"로그인 후에 서비스를 이용해주세요."
login_manager.login_message_category = "info"
@login_manager.user_loader
def load_user(id):
    return db.session.get(User, id)  # primary_key

def hash_phone_number(phone_number):
    # 휴대폰 번호를 SHA-256으로 해시
    sha256 = hashlib.sha256()
    sha256.update(phone_number.encode('utf-8'))
    hashed_phone = sha256.hexdigest()
    return hashed_phone

@app.after_request
def remove_header(response):
    response.headers['Server:'] = ''

    return response
#===================================================================================
# 데이터 삭제 작업 함수
def delete_old_records():
    with app.app_context():
        now = datetime.now()
        one_year_ago = now - timedelta(minutes=365)
        
        old_records = Privacy.query.filter(Privacy.visit_date <= one_year_ago).all()
        for record in old_records:
            db.session.delete(record)
        
        db.session.commit()


scheduler = BackgroundScheduler(daemon=True)
scheduler.add_job(delete_old_records, trigger='interval', days=1)
scheduler.start()

@app.route('/')
@app.route('/index')
@login_required
def index():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if current_user.attempts == "attempts_password":
        flash('비밀번호를 변경 후 서비스 이용이 가능합니다.')
        return redirect('authenticated')
    
    # 타임 스탬프
    today_weekday = datetime.now().weekday()
    weekdays = {0: "월요일", 1: "화요일", 2: "수요일", 3: "목요일", 4: "금요일", 5: "토요일", 6: "일요일"}
    weekday = weekdays.get(today_weekday, "")
    current_date = datetime.now().strftime('%Y년 %m월 %d일 ') + weekday
    current_time = datetime.now().strftime('\n%p %H:%M:%S')
    time = [ current_date, current_time]

    # 승인된 방문객 Sort_Desc
    approve_visitors = Visitor.query.filter_by(approve=1).order_by(Visitor.id.desc())
    for visitor in approve_visitors:
            visitor.department = aes.decrypt(visitor.department)
            visitor.phone = aes.decrypt(visitor.phone)

    # 출입 카드 목록
    card_list = Card.query.all()

    if approve_visitors:
        # 실시간 출입 방문객
        in_visitor = Visitor.query.filter_by(exit=0)
        in_visitor_card_none = Visitor.query.filter_by(exit=0, card_id=None)

        # 실시간 출입 방문객 카운팅
        in_visitor = in_visitor.count()
        in_visitor_card_none = in_visitor_card_none.count()
        in_visitor = in_visitor - in_visitor_card_none

        today = date.today()
        total_visitors = db.session.query(func.sum(Year.count)).scalar() # 총 방문객
        year = Year.query.filter_by(year=today.year).first() # 연간 방문객
        month = Month.query.filter_by(year=today.year, month=today.month).first() # 월간 방문객
        day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first() # 일간 방문객
        if year:
            yearly_visitor = year.count
            if month:
                monthly_visitor = month.count
                if day:
                    daily_visitor = day.count
                    visitor_count = [in_visitor, yearly_visitor, monthly_visitor, daily_visitor]
                else:
                    visitor_count = [in_visitor, yearly_visitor, monthly_visitor, 0]
            else:
                visitor_count = [in_visitor, yearly_visitor, 0, 0]
        else:
            visitor_count = [in_visitor, 0, 0, 0]
    
    # print('현재 로그인한 사용자: ' + str(current_user))
    return render_template('index.html', current_user=current_user, approve_visitors=approve_visitors, visitor_count=visitor_count, time=time, card_list=card_list, total_visitors=total_visitors)

#===================================================================================


#===================================================================================
admin_email, admin_password, admin2_email, admin2_password = account_list()
@app.route('/admin', methods=['GET', 'POST'])
def admin_page():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST':
        admin = User.query.filter_by(username='관리자').first()
        if admin:
            flash('이미 관리자 계정이 존재합니다.')
            return redirect('admin')
        username = '관리자'
        email = admin_email
        hashed_password = bcrypt.generate_password_hash(admin_password)
        admin = User(username, email, hashed_password, 'Admin', 'M', hashed_password, current_timestamp, current_timestamp, "관리자", "", "", request.remote_addr)
        db.session.add(admin)
        db.session.commit()
        flash("관리자 계정이 생성되었습니다.")
        return redirect('login')
    return render_template('admin.html')

@app.route('/admin_2', methods=['GET', 'POST'])
def admin2_page():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST':
        admin = User.query.filter_by(username='상황실').first()
        if admin:
            flash('이미 상황실 계정이 존재합니다.')
            return redirect('admin')
        username = '상황실'
        email = admin2_email
        hashed_password = bcrypt.generate_password_hash(admin2_password)
        admin = User(username, email, hashed_password, 'Admin', 'S', hashed_password, current_timestamp, current_timestamp, "일반", "", "", request.remote_addr)
        db.session.add(admin)
        db.session.commit()
        flash("상황실 계정이 생성되었습니다.")
        return redirect('login')

#===================================================================================

#===================================================================================

# 차트 관리 페이지
@app.route('/charts', methods=['GET', 'POST'])
@login_required
def visualization_chart():
    today = date.today()
    total_visitors = db.session.query(func.sum(Year.count)).scalar() # 총 방문객
    year = Year.query.filter_by(year=today.year).first() # 연간 방문객
    month = Month.query.filter_by(year=today.year, month=today.month).first() # 월간 방문객
    day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first() # 일간 방문객
    if year:
        yearly_visitor = year.count
        if month:
            monthly_visitor = month.count
            if day:
                daily_visitor = day.count
                visitor_count = [yearly_visitor, monthly_visitor, daily_visitor]
            else:
                visitor_count = [yearly_visitor, monthly_visitor, 0]
        else:
            visitor_count = [yearly_visitor, 0, 0]
    else:
        visitor_count = [0, 0, 0]

    daily_visitors = Day.query.filter_by(month=datetime.now().strftime('%m')).all()
    monthly_visitors = Month.query.filter_by(year=datetime.now().strftime('%Y')).all()
    yearly_visitors = Year.query.all()

    # 일간 방문자 수
    day_count = []
    what_day = []
    what_month = []

    # 월간 방문자 수
    month_count = []
    what_month_2 = []

    # 연간 방문자 수
    what_year = []
    year_count = []
    
    for daily_visitor in daily_visitors:
        day_count.append(daily_visitor.count)
        what_day.append(daily_visitor.day)
        what_month.append(daily_visitor.month)
    daily = [what_month, what_day, day_count]

    for monthly_visitor in monthly_visitors:
        what_month_2.append(monthly_visitor.month)
        month_count.append(monthly_visitor.count)
    monthly = [what_month_2, month_count]

    for yearly_visitor in yearly_visitors:
        what_year.append(yearly_visitor.year)
        year_count.append(yearly_visitor.count)
    yearly = [what_year, year_count]
    print(daily, monthly, yearly)
    return render_template('charts.html', daily=daily, monthly=monthly, yearly=yearly, visitor_count=visitor_count, total_visitors=total_visitors)

#===================================================================================


#===================================================================================

# 방문객 관리 페이지
@app.route('/manage_visitors', methods=['GET', 'POST'])
@login_required
def manage_visitors():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST' and current_user.rank == 'M':
        update_id = request.form['inputUpdateNumber']
        name = request.form['inputName']
        department = request.form['inputDepartment']
        phone = request.form['inputPhoneNumber']
        manager = request.form['inputManager']
        location = request.form.get('inputLocation')
        device = request.form.get('inputDevice')
        remarks = request.form.get('inputRemarks')
        object = request.form.get('inputObject')
        personal_computer = request.form.get('inputPC')
        work = request.form.get('inputWork')
        company_type = request.form.get('inputCompany')
        company = request.form.get('inputCompanyName')
        work_content = request.form.get('inputContent')
        model_name = request.form.get('inputModelName')
        serial_number = request.form.get('inputSerialNumber')
        reason = request.form.get('inputReason')
        work_division = request.form.get('inputWorkDivision')
        customer = request.form.get('inputCustomer')
        device_division = request.form.get('inputDeviceDivision')
        device_count = request.form.get('inputDeviceCount')

        update_visitor = Visitor.query.filter_by(id=update_id).first()
        update_visitor.name = name
        update_visitor.department = aes.encrypt(department)
        update_visitor.phone = aes.encrypt(phone)
        update_visitor.manager = manager

        if personal_computer == '1':
            update_visitor.personal_computer = True
        else:
            update_visitor.personal_computer = False

        update_visitor.model_name = model_name
        update_visitor.serial_number = serial_number
        update_visitor.pc_reason = reason

        if device == '1':
            update_visitor.device = True
        else:
            update_visitor.device = False

        update_visitor.customer = customer
        update_visitor.device_division = device_division
        update_visitor.device_count = device_count
        update_visitor.remarks = remarks
        update_visitor.location = location
        update_visitor.object = object

        if work == '1':
            update_visitor.work = True
        else:
            update_visitor.work = False

        update_visitor.work_division = work_division
        update_visitor.company_type = company_type
        update_visitor.company = company
        update_visitor.work_content = work_content
        # 수정하기 시 로그 남기기
        privacy_log = Privacy_log("수정", current_user.id, request.remote_addr, current_timestamp, "내방객 수정", update_visitor.name)
        db.session.add(privacy_log)
        db.session.commit()
        return redirect('manage_visitors')
    elif current_user.rank == 'M':

        # 타임 스탬프
        today_weekday = datetime.now().weekday()
        weekdays = {0: "월요일", 1: "화요일", 2: "수요일", 3: "목요일", 4: "금요일", 5: "토요일", 6: "일요일"}
        weekday = weekdays.get(today_weekday, "")
        current_date = datetime.now().strftime('%Y년 %m월 %d일 ') + weekday
        current_time = datetime.now().strftime('\n%p %H:%M:%S')
        time = [ current_date, current_time]

        # 승인된 방문객 Sort_Desc
        approve_visitors = Visitor.query.filter_by(approve=1).order_by(Visitor.id.desc())
        for visitor in approve_visitors:
            visitor.department = aes.decrypt(visitor.department)
            visitor.phone = aes.decrypt(visitor.phone)

        # 출입 카드 목록
        card_list = Card.query.all()
        # 랙 키 목록
        rack_key_list = Rack.query.all()
        if approve_visitors:
            # 실시간 출입 방문객
            in_visitor = Visitor.query.filter_by(exit=0)
            in_visitor_card_none = Visitor.query.filter_by(exit=0, card_id=None)

            # 실시간 출입 방문객 카운팅
            in_visitor = in_visitor.count()
            in_visitor_card_none = in_visitor_card_none.count()
            in_visitor = in_visitor - in_visitor_card_none

            today = date.today()
            total_visitors = db.session.query(func.sum(Year.count)).scalar() # 총 방문객
            year = Year.query.filter_by(year=today.year).first() # 연간 방문객
            month = Month.query.filter_by(year=today.year, month=today.month).first() # 월간 방문객
            day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first() # 일간 방문객
            if year:
                yearly_visitor = year.count
                if month:
                    monthly_visitor = month.count
                    if day:
                        daily_visitor = day.count
                        visitor_count = [in_visitor, yearly_visitor, monthly_visitor, daily_visitor]
                    else:
                        visitor_count = [in_visitor, yearly_visitor, monthly_visitor, 0]
                else:
                    visitor_count = [in_visitor, yearly_visitor, 0, 0]
            else:
                visitor_count = [in_visitor, 0, 0, 0]

        # 내방객 등록 - 부서 목록
        department_lists = Department.query.filter_by(user_id=current_user.id).all()
        return render_template('visitor_update.html', current_user=current_user, approve_visitors=approve_visitors, visitor_count=visitor_count, time=time, card_list=card_list, department_lists=department_lists, total_visitors=total_visitors, rack_key_list=rack_key_list)
    else:
        return render_template('404.html')
#===================================================================================

@app.route('/rack_visitors', methods=['GET','POST'])
@login_required
def rack_visitors():
    if request.method == 'GET' and current_user.rank == 'S':
        # 타임 스탬프
        today_weekday = datetime.now().weekday()
        weekdays = {0: "월요일", 1: "화요일", 2: "수요일", 3: "목요일", 4: "금요일", 5: "토요일", 6: "일요일"}
        weekday = weekdays.get(today_weekday, "")
        current_date = datetime.now().strftime('%Y년 %m월 %d일 ') + weekday
        current_time = datetime.now().strftime('\n%p %H:%M:%S')
        time = [current_date, current_time]

        # 승인된 방문객 Sort_Desc
        approve_visitors = Visitor.query.filter_by(approve=1).order_by(Visitor.id.desc())
        for visitor in approve_visitors:
            visitor.department = aes.decrypt(visitor.department)
            visitor.phone = aes.decrypt(visitor.phone)

        # 랙 키 목록
        rack_key_list = Rack.query.all()

        if approve_visitors:
            # 실시간 출입 방문객
            in_visitor = Visitor.query.filter_by(exit=0)
            in_visitor_card_none = Visitor.query.filter_by(exit=0, card_id=None)

            # 실시간 출입 방문객 카운팅
            in_visitor = in_visitor.count()
            in_visitor_card_none = in_visitor_card_none.count()
            in_visitor = in_visitor - in_visitor_card_none

            today = date.today()
            total_visitors = db.session.query(func.sum(Year.count)).scalar() # 총 방문객
            year = Year.query.filter_by(year=today.year).first() # 연간 방문객
            month = Month.query.filter_by(year=today.year, month=today.month).first() # 월간 방문객
            day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first() # 일간 방문객
            if year:
                yearly_visitor = year.count
                if month:
                    monthly_visitor = month.count
                    if day:
                        daily_visitor = day.count
                        visitor_count = [in_visitor, yearly_visitor, monthly_visitor, daily_visitor]
                    else:
                        visitor_count = [in_visitor, yearly_visitor, monthly_visitor, 0]
                else:
                    visitor_count = [in_visitor, yearly_visitor, 0, 0]
            else:
                visitor_count = [in_visitor, 0, 0, 0]
        
        return render_template('visitor_emergency.html', current_user=current_user, approve_visitors=approve_visitors, visitor_count=visitor_count, time=time, rack_key_list=rack_key_list, total_visitors=total_visitors)
    else:
        return render_template('404.html')

#===================================================================================

# 부서 관리 페이지
@app.route('/departments', methods=['GET', 'POST'])
@login_required
def manage_departments():
    if request.method == 'POST':
        department_type = request.form['select_department_type']
        department_name = request.form['add_department_name_value']
        if Department.query.filter_by(user_id=current_user.id, department_name=department_name).first():
            flash('이미 부서가 존재합니다.')
            return redirect('departments')
        department = Department(department_type, department_name, current_user.id)
        db.session.add(department)
        db.session.commit()
        return redirect('departments')
        return redirect('departments')
    else:
        departments = Department.query.filter_by(user_id=current_user.id).all()
        department_subsidiary = Department.query.filter_by(user_id=current_user.id, department_type="계열사").count()
        department_partner = Department.query.filter_by(user_id=current_user.id, department_type="협력사").count()
        department_bp = Department.query.filter_by(user_id=current_user.id, department_type="BP").count()
        department_etc = Department.query.filter_by(user_id=current_user.id, department_type="기타").count()
        department_counts = [department_subsidiary, department_partner, department_bp, department_etc]
        return render_template('manage_department.html', departments=departments, department_counts=department_counts)

# 부서 삭제 api
@app.route('/api/ajax_department_delete', methods=['POST'])
@login_required
def ajax_department_delete():
    data = request.get_json()
    print(data['delete_id'])
    delete_id = data['delete_id']
    department = Department.query.filter_by(id=delete_id, user_id=current_user.id).first()
    db.session.delete(department)
    db.session.commit()
    return jsonify()

# 기본 부서 생성
@app.route('/api/ajax_department_basic_create', methods=['POST'])
@login_required
def ajax_department_basic_create():
    # 출입 카드 DB Content 생성
    categories = ['CJ 올리브네트웍스', 'CJ(주)', 'CJ CGV', 'CJ 올리브영', 'CJ 프레시웨이', 'CJ 대한통운', 'CJ ENM', '디아이웨어']
    for category in categories:
        department = Department('계열사', category, current_user.id)
        db.session.add(department)
        db.session.commit()
    return jsonify(result = "success")

#===================================================================================


#===================================================================================

# 방문객 수정 api
@app.route('/visit_update', methods=['GET', 'POST'])
@login_required
def visit_update():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST':
        update_id = request.form['inputUpdateNumber']
        name = request.form['inputUpdateName']
        department = request.form['inputUpdateDepartment']
        phone = request.form['inputUpdatePhoneNumber']
        manager = request.form['inputUpdateManager']
        location = request.form.get('inputUpdateLocation')
        device = request.form.get('inputUpdateDevice')
        remarks = request.form.get('inputUpdateRemarks')
        object = request.form.get('inputUpdateObject')
        personal_computer = request.form.get('inputUpdatePC')
        work = request.form.get('inputUpdateWork')
        company_type = request.form.get('inputUpdateCompany')
        company = request.form.get('inputUpdateCompanyName')
        work_content = request.form.get('inputUpdateContent')
        model_name = request.form.get('inputUpdateModelName')
        serial_number = request.form.get('inputUpdateSerialNumber')
        reason = request.form.get('inputUpdateReason')
        work_division = request.form.get('inputUpdateWorkDivision')
        customer = request.form.get('inputUpdateCustomer')
        device_division = request.form.get('inputUpdateDeviceDivision')
        device_count = request.form.get('inputUpdateDeviceCount')

        update_visitor = Visitor.query.filter_by(id=update_id).first()
        update_visitor.name = name
        update_visitor.department = aes.encrypt(department)
        update_visitor.phone = aes.encrypt(phone)
        update_visitor.manager = manager

        if personal_computer == '1':
            update_visitor.personal_computer = True
        else:
            update_visitor.personal_computer = False

        update_visitor.model_name = model_name
        update_visitor.serial_number = serial_number
        update_visitor.pc_reason = reason

        if device == '1':
            update_visitor.device = True
        else:
            update_visitor.device = False
        
        update_visitor.customer = customer
        update_visitor.device_division = device_division
        update_visitor.device_count = device_count
        update_visitor.remarks = remarks
        update_visitor.location = location
        update_visitor.object = object

        if work == '1':
            update_visitor.work = True
        else:
            update_visitor.work = False

        update_visitor.work_division = work_division
        update_visitor.company_type = company_type
        update_visitor.company = company
        update_visitor.work_content = work_content

        task_change = Privacy_log("수정", current_user.id, request.remote_addr, current_timestamp, "내방객 수정", update_visitor.name)
        db.session.add(task_change)

        db.session.commit()
        return redirect('visit')

#===================================================================================


#===================================================================================

# 카드 관리 페이지
@app.route('/manage_cards', methods=['GET', 'POST'])
@login_required
def manage_cards():
    if current_user.rank == 'M':
        cards = db.session.query(Card.card_type).distinct().all()
        categories = []
        for card in cards:
            categories.append(card[0])

        card_counts = {}
        recall_cards = []
        use_cards = []
        for category in categories:
            card_counts[f'{category}'] = Card.query.filter_by(card_type=category).order_by(func.cast(Card.card_num, Integer).asc()).all()
            # 회수된 카드
            recall_cards.append(Card.query.filter_by(card_type=category, card_status='회수').count())

        # 총 출입 카드
        total_cards = Card.query.count()

        return render_template('manage_cards.html', card_counts=card_counts, recall_cards=recall_cards, total_cards=total_cards)
    else:
        return render_template('404.html')

#===================================================================================


#===================================================================================

# 랙키 관리 페이지
@app.route('/manage_rack_keys', methods=['GET', 'POST'])
@login_required
def manage_rack_keys():
    if current_user.rank == 'M' or current_user.rank == 'S':
        keys = db.session.query(Rack.key_type).distinct().all()
        categories = []
        for key in keys:
            categories.append(key[0])

        key_counts = {}
        recall_keys = []
        use_keys = []
        for category in categories:
            key_counts[f'{category}'] = Rack.query.filter_by(key_type=category).order_by(func.cast(Rack.key_num, Integer).asc()).all()
            # 회수된 키
            recall_keys.append(Rack.query.filter_by(key_type=category, key_status='회수').count())

        # 총 출입 키
        total_keys = Rack.query.count()
        
        return render_template('manage_rack_key.html', key_counts=key_counts, recall_keys=recall_keys, total_keys=total_keys)
    else:
        return render_template('404.html')

#===================================================================================


#===================================================================================

# 로그 관리 페이지
@app.route('/manage_logs', methods=['GET', 'POST'])
@login_required
def manage_logs():
    if current_user.rank == 'M':
        user_log = User_log.query.all()
        return render_template('manage_logs.html', user_log=user_log)
    else:
        return render_template('404.html')

@app.route('/user_logs', methods=['GET', 'POST'])
@login_required
def user_logs():
    rank_list = ['G1','G2','G3','G4','G5','G6','G7','S','임원']
    if current_user.rank in rank_list:
        user_log = User_log.query.filter_by(user_id=current_user.id).all()
        return render_template('user_logs.html', user_log=user_log)
    else:
        return render_template('404.html')

@app.route('/account_logs', methods=['GET', 'POST'])
@login_required
def account_logs():
    if current_user.rank == 'M':
        user = User.query.filter_by().order_by(User.id).all()
        remove_user = Account_log.query.all()
        print(user)
        return render_template('account_logs.html', user=user, remove_user=remove_user)
    else:
        return render_template('404.html')

@app.route('/permission', methods=['GET', 'POST'])
@login_required
def permission():
    if current_user.rank == 'M':
        user = User.query.filter_by().order_by(User.id).all()
        permission = Permission_log.query.all()
        return render_template('permission.html', user=user, permission=permission)
    else:
        return render_template('404.html')

@app.route('/failure_logs', methods=['GET', 'POST'])
@login_required
def login_failure_logs():
    login_fail_log = Login_failure_log.query.filter_by(user_id=current_user.id).order_by(Login_failure_log.id.desc()).all()
    return render_template('login_failure_logs.html', login_fail_log=login_fail_log)

@app.route('/password_change_at', methods=['GET', 'POST'])
@login_required
def password_change_log():
    if current_user.rank == 'M':
        password_change_log = Password_change_log.query.order_by(Password_change_log.id.desc()).all()
        return render_template('password_change_at.html', password_change_log=password_change_log)
    else:
        password_change_log = Password_change_log.query.filter_by(user_id=current_user.id).order_by(Password_change_log.id.desc()).all()
        return render_template('password_change_at.html', password_change_log=password_change_log)

@app.route('/privacy_logs', methods=['GET', 'POST'])
@login_required
def privacy_logs():
    if current_user.rank == 'M':
        register_log = Privacy_log.query.filter_by(task_title='등록').all()
        approve_log = Privacy_log.query.filter_by(task_title='승인').all()
        reject_log = Privacy_log.query.filter_by(task_title='반려').all()
        change_log = Privacy_log.query.filter_by(task_title='수정').all()
        inquiry_log = Privacy_log.query.filter_by(task_title='조회').all()
        delete_log= Privacy_log.query.filter_by(task_title='삭제').all()
        return render_template('privacy_logs.html', register_log=register_log, approve_log=approve_log, reject_log=reject_log, change_log=change_log, inquiry_log=inquiry_log, delete_log=delete_log)
    else:
        return render_template('404.html')

@app.route('/api/ajax_delete_account', methods=['POST'])
@login_required
def ajax_delete_account():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = request.get_json()
    account_num = data['delete_btn']
    user = User.query.filter_by(id=account_num).first()
    if user == None:
        return "No Data"
    
    delete_user = Account_log(user.email, current_timestamp)
    db.session.add(delete_user)
    db.session.delete(user)
    db.session.commit()
    return jsonify()

@app.route('/api/ajax_permission_change', methods=['POST'])
@login_required
def ajax_permission_change():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = request.get_json()
    select_user = data['user']
    new_permission = data['permission']
    user = User.query.filter_by(id=select_user).first()
    if user == None:
        return "No Data"
    
    if user.permission == new_permission:
        return "Same Data"
    
    permission = Permission_log(user.email, user.permission, new_permission, current_timestamp)
    user.permission = new_permission
    db.session.add(permission)
    db.session.commit()
    return jsonify()

@app.route('/<pagename>')
def admin(pagename):
    return render_template(pagename+'.html')

#===================================================================================


#===================================================================================

def contains_consecutive(string, length):
    for i in range(len(string) - length + 1):
        if all(ord(string[i + j]) == ord(string[i]) + j for j in range(1, length)):
            return True
    return False

def contains_decreasing(string, length):
    for i in range(len(string) - length + 1):
        if all(ord(string[i + j]) == ord(string[i]) - j for j in range(1, length)):
            return True
    return False

def is_keyboard_consecutive(string, length):
    keyboard_rows = ['qwertyuiop', 'asdfghjkl', 'zxcvbnm', 'poiuytrewq', 'lkjhgfdsa', 'mnbvcxz', 'qazwsxedcrfvtgbyhnujmikolp', 'polikujmyhntgbrfvedcwsxqaz', 'zaqxswcdevfrbgtnhymjukilop', 'plokimjunhybgtvfrcdexswzaq','abcdefghijklmnopqrstuvwxyz', 'zyxwvutsrqponmlkjihgfedcba']
    for row in keyboard_rows:
        for i in range(len(row) - length + 1):
            if row[i:i+length] in string:
                return True
    return False

# 회원가입 개인정보 수집 동의
@app.route('/policy-register', methods=['GET','POST'])
def policy_register():
    return render_template('policy-register.html')

# 회원가입 페이지
@app.route('/register', methods=['GET', 'POST'])
def register():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST':
        # 폼으로부터 입력받은 데이터 가져오기
        username = request.form['username']
        email = request.form['email']
        email_fix = request.form['email_fix']
        password1 = request.form['password1']
        password2 = request.form['password2']
        department = request.form['registerDepartment']
        rank = request.form['registerRank']
        password_question = request.form['password_question']
        password_hint_answer = request.form['password_hint']
        
        email_new = email + email_fix
        print(email)
        print(email_fix)
        # 유효성 검사
        user = User.query.filter_by(email=email_new).first()
        if user:
            flash("이미 가입된 이메일입니다.")
        elif len(email) < 3:
            flash("이메일은 3자 이상이어야 합니다.")
        elif len(email) > 20:
            flash("이메일은 20자 이하여야 합니다.")
        elif re.search(r'(.)\1\1\1', email.lower()) or contains_consecutive(email, 4) or contains_decreasing(email, 4) or is_keyboard_consecutive(email.lower(), 4):
            flash("이메일에 4자 이상의 반복문자나 반복숫자를 사용할 수 없습니다.")
        elif len(username) < 2:
            flash("이름은 2자 이상이어야 합니다.")
        elif len(username) > 10:
            flash("이름은 10자 이하여야 합니다.")
        elif password1 != password2:
            flash("비밀번호와 비밀번호재입력이 서로 다릅니다.")
        elif len(password1) < 7:
            flash("비밀번호는 8자 이상이어야 합니다.")
        elif len(password1) > 14:
            flash("비밀번호는 14자 이하여야 합니다.")
        else:
            # 최소 3종류 이상 포함하는지 검사
            categories = 0
            if re.search(r'[A-Z]', password1):
                categories += 1
            if re.search(r'[a-z]', password1):
                categories += 1
            if re.search(r'\d', password1):
                categories += 1
            if re.search(r'[!@#$%^&*(),.?":{}|<>]', password1):
                categories += 1
            if categories < 4:
                flash('비밀번호 3종 복잡도를 만족하지 않습니다.')
            elif re.search(r'(.)\1{3,}', password1.lower()) or contains_consecutive(password1, 4) or contains_decreasing(password1, 4) or is_keyboard_consecutive(password1.lower(), 4):
                flash('4자 이상의 연속 문자를 사용할 수 없습니다.')
            else:
                # 비밀번호 암호화
                hashed_password = bcrypt.generate_password_hash(password1)
                email = email + email_fix
                user = User(username, email, hashed_password, department, rank, hashed_password, current_timestamp, current_timestamp, "일반", password_question, password_hint_answer, request.remote_addr)

                # 비밀번호 이력 보관
                db.session.add(user)
                db.session.commit()

                # 회원가입이 성공적으로 완료됨을 알리는 메시지 표시
                flash("회원가입 완료되었습니다.")
                return redirect('login')

    form_val = request.args.get('checkVal')
    if form_val != SECRET_TOKEN:
        return redirect('policy-register')
    # GET 요청인 경우 회원가입 양식을 표시
    return render_template('register.html')

def generate_random_verification_code(email_address, otp_create_time):
    # 6자리 랜덤 숫자 생성 (100000부터 999999 사이의 범위)
    otp = random.randint(100000, 999999)
    session[f'otp_{email_address}'] = otp  # 세션에 인증번호 저장
    session[f'time_{email_address}'] = otp_create_time  # 인증번호 생성 시간 저장
    return str(otp)

@app.route('/api/register_email_valid', methods=['POST'])
def register_email_valid():
    current_timestamp = datetime.now()
    data = request.get_json()
    email = data['email']
    
    user = User.query.filter_by(email=email).first() 
    if user:
        return "Use Email"
    elif len(email) < 3:
        return "Short"
    elif len(email) > 20:
        return "Long"
    else:
        random_auth_number = generate_random_verification_code(email, current_timestamp)
        email_subject = "[O`PASS] Send Auth Number"
        content_file_path = '/home/cjadmin/web/O_Pass/email_auth_number.txt'

        with open(content_file_path, 'r') as file:
            original_content = file.read()

        updated_content = '[' + random_auth_number + ']' + '\n\n' + original_content

        with open(content_file_path, 'w') as file:
            file.write(updated_content)

        command = (
            f"cat '{content_file_path}' | mail -s '{email_subject}' {email}"
        )
        os.system(command)
        
        with open(content_file_path, 'w') as file:
            file.write(original_content)
            
    return jsonify(result=random_auth_number)

@app.route('/api/register_auth_number_valid', methods=['POST'])
def register_auth_number_valid():
    # 현재 시간을 가져옵니다.
    current_timestamp = datetime.now()

    data = request.get_json()
    otp = data['otp']
    email_address = data['email']
    
    session_otp = session.get(f'otp_{email_address}')  # 세션에 저장된 인증번호 가져오기
    session_time = session.get(f'time_{email_address}')  # 세션에 저장된 생성 시간 가져오기
    
    # session_time을 offset-naïve datetime 객체로 만듭니다.
    session_time = session_time.replace(tzinfo=None)
    print('입력한 데이터')
    print(otp, type(otp))

    print('세션에서 받아온 데이터')
    print(session_otp, session_time, type(session_otp))

    different_time = current_timestamp - session_time

    # 세션에 인증번호와 생성 시간이 저장되어 있지 않은 경우
    if not session_otp or not session_time:
        return "No Data"

    # 시간이 만료된 경우
    if different_time.total_seconds() > 300:
    	# 세션에서 인증번호와 생성 시간을 삭제
        session.pop(f'otp_{email_address}')  
        session.pop(f'time_{email_address}')
        return "Time Out"
    
    # 검증에 성공한 경우
    if otp == str(session_otp):
    	# 세션에서 인증번호와 생성 시간을 삭제
        session.pop(f'otp_{email_address}')  
        session.pop(f'time_{email_address}')
        return jsonify(result="success")
    else:
        return "No Auth"

def check_password_change(user):
    if user.password_changed_at:
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        current_date = datetime.strptime(current_date, '%Y-%m-%d %H:%M:%S')
        delta = current_date - user.password_changed_at
        print("비밀번호 변경 시간:", user.password_changed_at)
        print("현재 시간:", current_date)
        print(delta)
        if delta > timedelta(days=180):  # 6개월 이상 경과한 경우
            print("비밀번호 변경이 필요합니다.")
            print("현재 초과시간:", delta)
            flash("마지막 비밀번호 변경일로부터 6개월이 경과하였습니다. 비밀번호를 변경해주세요.")
            print("알림 메시지:", get_flashed_messages())
        else:
            pass

LOGIN_BLOCK_DURATION = 30 # 로그인 제한 기간 (분)
MAX_LOGIN_ATTEMPTS = 5 # 로그인 실패 허용 횟수

# 로그인 기능
@app.route('/login', methods=['GET', 'POST'])
def login():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST':
        current_date = datetime.now()
        email = request.form['email']
        password = request.form['password1']

        user = User.query.filter_by(email=email).first()
        if user:
            if user.login_blocked_until and user.login_blocked_until > datetime.now():
                # 로그인이 제한된 경우
                block_remaining = (user.login_blocked_until - datetime.now()).seconds // 60
                flash('일시적으로 로그인이 제한되었습니다. 잠시 후 다시 시도해주세요. (제한 시간: {}분 남음)'.format(block_remaining))
                return render_template('login.html')
            
            if bcrypt.check_password_hash(user.password, password):
                # 로그인 성공
                session['user_id'] = user.username
                user.login_attempts = 0  # 로그인 시도 횟수 초기화
                user.login_blocked_until = None  # 로그인 제한 해제
                
                login_user(user)
                # 비인가 접근 및 오남용 등에 대한 경고 문구, 로그인 시 이전 로그인 정보 표시
                user_log = User_log(request.remote_addr, current_timestamp, user.id)

                session.permanent = True

                db.session.add(user_log)
                db.session.commit()

                current_login = User_log.query.filter_by(user_id=current_user.id).order_by(User_log.id.desc()).first()
                # check_password_change(user) # 비밀번호 변경 주기 체크
                if current_user.rank == 'M':
                    if current_login.login_timestamp:
                        flash('이전 로그인 일시: ' + str(current_login.login_timestamp) + " 접근 IP주소: " + current_login.ip_address)
                        if user.attempts == 'attempts_password':
                            return redirect('authenticated')
                    else:
                        flash('이전 로그인 일시: 없음'+ " 접근 IP주소: " + current_login.ip_address)
                        if user.attempts == 'attempts_password':
                            return redirect('authenticated')
                    return redirect(url_for('manage_visitors'))
                elif current_user.rank == 'S':
                    if current_login.login_timestamp:
                        flash('이전 로그인 일시: ' + str(current_login.login_timestamp) + " 접근 IP주소: " + current_login.ip_address)
                        if user.attempts == 'attempts_password':
                            return redirect('authenticated')
                    else:
                        flash('이전 로그인 일시: 없음'+ " 접근 IP주소: " + current_login.ip_address)
                        if user.attempts == 'attempts_password':
                            return redirect('authenticated')
                    return redirect(url_for('rack_visitors'))
                else:
                    if current_login.login_timestamp:
                        flash('이전 로그인 일시: ' + str(current_login.login_timestamp) + " 접근 IP주소: " + current_login.ip_address)
                        if user.attempts == 'attempts_password':
                            return redirect('authenticated')
                    else:
                        flash('이전 로그인 일시: 없음'+ " 접근 IP주소: " + current_login.ip_address)
                        if user.attempts == 'attempts_password':
                            return redirect('authenticated')
                    return redirect(url_for('index'))

            else:
                if user.login_blocked_until and user.login_blocked_until < datetime.now():
                    user.login_attempts = 0  # 로그인 시도 횟수 초기화
                    user.login_blocked_until = None  # 로그인 제한 해제
                    db.session.commit()

                # 로그인 실패
                user.login_attempts += 1  # 로그인 시도 횟수 증가
                if user.login_attempts >= MAX_LOGIN_ATTEMPTS:
                    # 일정 횟수 이상 실패 시 로그인 제한
                    user.login_blocked_until = datetime.now() + timedelta(minutes=LOGIN_BLOCK_DURATION)
                    block_remaining = LOGIN_BLOCK_DURATION

                    # 로그인 실패 로그 기록
                    login_fail = Login_failure_log(user.id, current_timestamp)
                    db.session.add(login_fail)
                    db.session.commit()

                    flash('5번 이상 실패할 경우 로그인이 제한됩니다.')
                else:
                    # 로그인 실패 로그 기록
                    login_fail = Login_failure_log(user.id, current_timestamp)
                    db.session.add(login_fail)
                    db.session.commit()

                    flash('5번 이상 실패할 경우 로그인이 제한됩니다.')
                db.session.commit()
                return render_template('login.html')
        else:
            flash('5번 이상 실패할 경우 로그인이 제한됩니다.')
            return render_template('login.html')
    else: # GET 로그인 페이지
        return render_template('login.html')

# 로그아웃 페이지
@app.route('/logout')
@login_required
def logout():
    session.pop('user_id', None)
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # 현재 로그인된 사용자 중 가장 마지막 접속 기록 ID 추출
    logout_log = User_log.query.filter_by(user_id=current_user.id).order_by(User_log.id.desc()).first()
    # 마지막 접속 ID의 로그아웃 스탬프 컬럼에 현재 시간 기록
    logout_log.logout_timestamp = current_timestamp
    # DB 적용
    db.session.commit()
    # 로그아웃 수행
    logout_user()
    print('logout success!')
    return redirect(url_for('login'))

#def random_string():
#    new_pw_len = 10
#    pw_candidate = string.ascii_letters
#    new_pw = ""
#    for i in range(new_pw_len):
#        new_pw += random.choice(pw_candidate)
#    print("\n생성된 랜덤 비밀번호", new_pw)
#    return new_pw
def random_string():
    new_pw_len = 10
    pw_candidate = string.ascii_letters + string.digits + string.punctuation
    new_pw = ""
    while True:
        new_pw = ''.join(random.choice(pw_candidate) for _ in range(new_pw_len))
        
        # 조건 검사
        if (len(new_pw) >= 8 and
            any(c.isdigit() for c in new_pw) and
            any(c.isalpha() for c in new_pw) and
            any(c in string.punctuation for c in new_pw) and
            not any(new_pw[i:i+4] == new_pw[i]*4 for i in range(len(new_pw)-3))):
            break
        
    print("\n생성된 랜덤 비밀번호:", new_pw)
    return new_pw


# 비밀번호 찾기 페이지
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    return render_template('forgot-password.html')

# 이메일 유효성 검사
@app.route('/api/password_forgot_email_valid', methods=['POST'])
def password_forgot_email_valid():
    data = request.get_json()
    email = data['email']
    user = User.query.filter_by(email=email).first()
    
    if user:
        return jsonify(result = "success")
    else:
        return "No Email"
    
# 패스워드 정답 유효성 검사
# 패스워드 정답 유효성 검사
@app.route('/api/password_forgot_answer_valid', methods=['POST'])
def password_forgot_answer_valid():
    data = request.get_json()
    email = data['email']
    question = data['question']
    answer = data['answer']
    user = User.query.filter_by(email=email, password_question=question, password_hint_answer=answer).first()
    
    if user:
        connect_to_database()
        hash_password = random_string()
        user.password = bcrypt.generate_password_hash(hash_password)
        user.attempts = "attempts_password"
        db.session.commit()

        email_subject = "[O`PASS] Send Password"
        content_file_path = '/home/cjadmin/web/O_Pass/email_content.txt'

        # 원본 파일 내용 읽기
        with open(content_file_path, 'r') as file:
            original_content = file.read()

        # hash_password 값을 파일 내용 맨 앞에 추가
        updated_content = '[ ' + hash_password + ' ]' + '\n\n' + original_content

        # 수정된 내용을 파일에 다시 쓰기
        with open(content_file_path, 'w') as file:
            file.write(updated_content)

        # 명령어 문자열 생성
        command = (
            f"cat '{content_file_path}' | mail -s '{email_subject}' {email}"
        )
        os.system(command)

        # 다시 원본으로 교체
        with open(content_file_path, 'w') as file:
            file.write(original_content)

        return jsonify(result = "success")
    else:
        return "No Answer"

#===================================================================================


#===================================================================================

# 내방객 등록 페이지
@app.route('/visit', methods=['GET', 'POST'])
@login_required
def visitor():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if request.method == 'POST':
        name = request.form['inputName']
        department = request.form['inputDepartment']
        object = request.form['inputObject']
        phone = request.form['inputPhoneNumber']
        manager = request.form['inputManager']
        created_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        personal_computer = request.form.get('inputPC')
        device = request.form.get('inputDevice')
        remarks = request.form.get('inputRemarks')
        work = request.form.get('inputWork')

        if personal_computer:
            personal_computer = True
            model_name = request.form.get('inputModelName')
            serial_number = request.form.get('inputSerialNumber')
            reason = request.form.get('inputReason')
        else:
            personal_computer = False
            model_name = None
            serial_number = None
            reason = None

        if device:
            device = True
            customer = request.form.get('inputCustomer')
            device_division = request.form.get('inputDeviceDivision')
            device_count = request.form.get('inputDeviceCount')
            remarks = request.form.get('inputRemarks')
        else:
            device = False
            customer = None
            device_division = None
            device_count = None
            remarks = None

        if work:
            work = True
            location = request.form.get('inputLocation')
            work_division = request.form.get('inputWorkDivision')
            company_type = request.form.get('inputCompany')
            company_name = request.form.get('inputCompanyName')
            work_content = request.form.get('inputContent')
        else:
            work = False
            location = None
            work_division = None
            company_type = None
            company_name = None
            work_content = None

            # aes.encrypt(department)
            # aes.encrypt(phone)

        # 내방객 등록하기 - 이름, 부서, 번호, 작업위치, 담당자, 장비체크, 비고, 방문목적, 등록시간, 승인, 사전/현장, 작업체크, 회사종류, 회사이름, 작업내용
        visitor = Visitor(name, aes.encrypt(department), aes.encrypt(phone), location, manager, device, remarks, object, created_time, 0, "사전 등록", work, company_type, company_name, work_content, current_user.id, personal_computer, model_name, serial_number, reason, work_division, customer, device_division, device_count)
        task_change = Privacy_log("등록", current_user.id, request.remote_addr, current_timestamp, "내방객 등록", name)
        db.session.add(visitor)
        db.session.add(task_change)
        db.session.commit()
        return redirect(url_for('visitor'))
    else:
        # GET - 승인되지 않은 방문객 정보
        if current_user.permission == '관리자':
            visitor_info = Visitor.query.filter_by(approve=0)
            for visitor in visitor_info:
                visitor.department = aes.decrypt(visitor.department)
                visitor.phone = aes.decrypt(visitor.phone)

            # 내방객 등록 - 부서 목록
            department_lists = Department.query.filter_by(user_id=current_user.id).all()
            print(department_lists)

            return render_template('visitor.html', department_lists=department_lists, visitor_info=visitor_info)
        else:
            visitor_info = Visitor.query.filter_by(approve=0, writer=current_user.id)
            for visitor in visitor_info:
                visitor.department = aes.decrypt(visitor.department)
                visitor.phone = aes.decrypt(visitor.phone)

            # 내방객 등록 - 부서 목록
            department_lists = Department.query.filter_by(user_id=current_user.id).all()
            print(department_lists)

            return render_template('visitor.html', department_lists=department_lists, visitor_info=visitor_info)



# 승인 버튼 클릭시 로직 ajax
@app.route('/api/ajax_approve', methods=['POST'])
@login_required
def ajax_approve():
    data = request.get_json()
    print(data['visitor_id'])
    print(data['approve'])

    visitor = Visitor.query.filter_by(id=data['visitor_id']).first()
    visitor.approve = 1
    visitor.exit = 0
    visitor.approve_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    visitor.approve_log = current_user.id

    today = date.today()
    year = Year.query.filter_by(year=today.year).first()
    if not year:
        year = Year(year=today.year, count=1)
        db.session.add(year)
    else:
        year.count += 1

    month = Month.query.filter_by(year=today.year, month=today.month).first()
    if not month:
        month = Month(year=today.year, month=today.month, count=1)
        db.session.add(month)
    else:
        month.count += 1

    day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first()
    if not day:
        day = Day(year=today.year, month=today.month, day=today.day, count=1)
        db.session.add(day)
    else:
        day.count += 1

    if visitor.personal_computer == 0:
        personal_computer = "미반입"
    else:
        personal_computer = "반입"
    if visitor.device == 0:
        device = "미반입"
        privacy_device = False
    else:
        device = "반입"
        privacy_device = True

    if visitor.work == 0:
        work = "해당 없음"
        privacy_work = False
    else:
        work = "해당"
        privacy_work = True

    print(visitor.id)
    # qr_expired_date = datetime.now().strftime('%Y-%m-%d')
    # qr_data = "https://opass.cj.net/?id="
    # qr_img = qrcode.make(qr_data)
    # save_path = 'qrcode.png'
    # qr_img.save(save_path)

    privacy_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if visitor.registry == "사전 등록":
        image_send_sms_previous(visitor.name, visitor.approve_date, visitor.object, visitor.location, visitor.manager, aes.decrypt(visitor.phone), device, work, visitor.company, visitor.work_content, personal_computer, visitor.model_name, visitor.work_division, visitor.device_division, visitor.device_count)
        privacy = Privacy(visitor.name, aes.encrypt(visitor.department), aes.encrypt(visitor.phone), visitor.manager, privacy_device, privacy_work, visitor.remarks, visitor.object, visitor.location, visitor.company_type, visitor.company, visitor.work_content, privacy_date, "사전 등록", visitor.personal_computer, visitor.model_name, visitor.serial_number, visitor.pc_reason, visitor.work_division, visitor.customer, visitor.device_division, visitor.device_count)
    else:
        image_send_sms_current(visitor.name, visitor.approve_date, visitor.object, visitor.location, visitor.manager, aes.decrypt(visitor.phone), device, work, visitor.company, visitor.work_content, personal_computer, visitor.model_name, visitor.work_division, visitor.device_division, visitor.device_count)
        privacy = Privacy(visitor.name, aes.encrypt(visitor.department), aes.encrypt(visitor.phone), visitor.manager, privacy_device, privacy_work, visitor.remarks, visitor.object, visitor.location, "", visitor.company, visitor.work_content, privacy_date, "현장 등록", visitor.personal_computer, visitor.model_name, visitor.serial_number, visitor.pc_reason, visitor.work_division, visitor.customer, visitor.device_division, visitor.device_count)
    task_change = Privacy_log("승인", current_user.id, request.remote_addr, privacy_date, "내방객 승인", visitor.name)

    db.session.add(task_change)

    db.session.add(privacy)
    db.session.commit()
    return jsonify(result = "success")

# 반려 버튼 클릭시 로직 ajax
@app.route('/api/ajax_deny', methods=['POST'])
@login_required
def ajax_deny():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = request.get_json()
    print(data['visitor_id'])
    print(data['approve'])
    visitor = Visitor.query.filter_by(id=data['visitor_id']).first()

    task_change = Privacy_log("반려", current_user.id, request.remote_addr, current_timestamp, "내방객 반려", visitor.name)
    db.session.add(task_change)

    db.session.delete(visitor)
    db.session.commit()
    return jsonify(result = "success")
# 긴급 승인 버튼 클릭시 로직 ajax
# @app.route('/api/ajax_emergency_approve', methods=['POST'])
# def ajax_emergency_approve():
#     data = request.get_json()
#     print(data['visitor_id'])
#     print(data['approve'])

#     visitor = Visitor.query.filter_by(id=data['visitor_id']).first()
#     visitor.approve = 1
#     visitor.exit = 0

#     today = date.today()
#     year = Year.query.filter_by(year=today.year).first()
#     if not year:
#         year = Year(year=today.year, count=1)
#         db.session.add(year)

#     month = Month.query.filter_by(year=today.year, month=today.month).first()
#     if not month:
#         month = Month(year=today.year, month=today.month, count=1)
#         db.session.add(month)

#     day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first()
#     if not day:
#         day = Day(year=today.year, month=today.month, day=today.day, count=1)
#         db.session.add(day)
#     else:
#         year.count += 1
#         month.count += 1
#         day.count += 1

#     db.session.commit()
#     return jsonify(result = "success")

#===================================================================================


#===================================================================================

# 내방객 관리 페이지 -  퇴실 버튼 클릭시 로직 ajax
@app.route('/api/ajax_exit', methods=['POST'])
@login_required
def ajax_exit():
    data = request.get_json()
    visitor = Visitor.query.filter_by(id=data['exit_id']).first()

    if visitor.card_id is None: # 카드를 안 받은 사람은 퇴실 X
        return "Card None"

    if visitor.exit_date == None and visitor.exit == 0:
        visitor.exit_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        visitor.exit = 1
        visitor.card.card_status = "회수"
        visitor.card_id = None
        if visitor.work == 1 and visitor.rack_id:
            visitor.rack.key_status = "회수"
            visitor.rack_id = None
        visitor.exit_log = current_user.id
        db.session.commit()
    else:
        return "Exit Error"

    return jsonify(response = "success")

# 내방객 관리 페이지 - 퇴실 갱신 로직 ajax
@app.route('/api/ajax_re_exit', methods=['POST'])
@login_required
def ajax_re_exit():
    data = request.get_json()
    visitor = Visitor.query.filter_by(id=data['exit_id']).first()

    if visitor.exit == 1:
        visitor.exit_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        visitor.exit_log = current_user.id
        db.session.commit()
    else:
        return "Exit Error"

    return jsonify(response = "success")

# 내방객 관리 페이지 - 체크 박스 퇴실 api
@app.route('/api/ajax_index_exit_checkbox', methods=['POST'])
@login_required
def ajax_index_exit_checkbox():
    data = request.get_json()
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        if visitor.card_id is None:
            return "No Card"

        if visitor.exit_date == None and visitor.exit == 0:
            visitor.exit_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            visitor.exit = 1
            visitor.card.card_status = "회수"
            visitor.card_id = None
            if visitor.work == 1 and visitor.rack_id:
                visitor.rack.key_status = "회수"
                visitor.rack_id = None
            visitor.exit_log = current_user.id
            db.session.commit()
        else:
            return "Exited"
    return jsonify(result = "success")

# 내방객 등록 페이지 - 체크 박스 승인 api
@app.route('/api/ajax_visit_approve_checkbox', methods=['POST'])
@login_required
def ajax_visit_approve_checkbox():
    data = request.get_json()
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        visitor.approve = 1
        visitor.exit = 0
        visitor.approve_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        visitor.approve_log = current_user.id

        today = date.today()
        year = Year.query.filter_by(year=today.year).first()
        if not year:
            year = Year(year=today.year, count=1)
            db.session.add(year)
        else:
            year.count += 1

        month = Month.query.filter_by(year=today.year, month=today.month).first()
        if not month:
            month = Month(year=today.year, month=today.month, count=1)
            db.session.add(month)
        else:
            month.count += 1

        day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first()
        if not day:
            day = Day(year=today.year, month=today.month, day=today.day, count=1)
            db.session.add(day)
        else:
            day.count += 1

        if visitor.personal_computer == 0:
            personal_computer = "미반입"
        else:
            personal_computer = "반입"

        if visitor.device == 0:
            device = "미반입"
            privacy_device = False
        else:
            device = "반입"
            privacy_device = True

        if visitor.work == 0:
            work = "해당 없음"
            privacy_work = False
        else:
            work = "해당"
            privacy_work = True

        privacy_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if visitor.registry == "사전 등록":
            image_send_sms_previous(visitor.name, visitor.approve_date, visitor.object, visitor.location, visitor.manager, aes.decrypt(visitor.phone), device, work, visitor.company, visitor.work_content, personal_computer, visitor.model_name, visitor.work_division, visitor.device_division, visitor.device_count)
            privacy = Privacy(visitor.name, aes.encrypt(visitor.department), aes.encrypt(visitor.phone), visitor.manager, privacy_device, privacy_work, visitor.remarks, visitor.object, visitor.location, visitor.company_type, visitor.company, visitor.work_content, privacy_date, "사전 등록", visitor.personal_computer, visitor.model_name, visitor.serial_number, visitor.pc_reason, visitor.work_division, visitor.customer, visitor.device_division, visitor.device_count)
        elif visitor.registry == "현장 등록":
            image_send_sms_current(visitor.name, visitor.approve_date, visitor.object, visitor.location, visitor.manager, aes.decrypt(visitor.phone), device, work, visitor.company, visitor.work_content, personal_computer, visitor.model_name, visitor.work_division, visitor.device_division, visitor.device_count)
            privacy = Privacy(visitor.name, aes.encrypt(visitor.department), aes.encrypt(visitor.phone), visitor.manager, privacy_device, privacy_work, visitor.remarks, visitor.object, visitor.location, "", visitor.company, visitor.work_content, privacy_date, "현장 등록", visitor.personal_computer, visitor.model_name, visitor.serial_number, visitor.pc_reason, visitor.work_division, visitor.customer, visitor.device_division, visitor.device_count)
        task_change = Privacy_log("승인", current_user.id, request.remote_addr, privacy_date, "내방객 승인", visitor.name)
        db.session.add(task_change)

        db.session.add(privacy)
        db.session.commit()
    return jsonify(result = "success")

# 내방객 등록 페이지 - 체크 박스 반려 api
@app.route('/api/ajax_visit_deny_checkbox', methods=['POST'])
@login_required
def ajax_visit_deny_checkbox():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = request.get_json()
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()

        task_change = Privacy_log("반려", current_user.id, request.remote_addr, current_timestamp, "내방객 반려", visitor.name)
        db.session.add(task_change)

        db.session.delete(visitor)
        db.session.commit()
        
    return jsonify(result = "success")

# visit 체크 박스 긴급 승인 api - G6에게 이메일, 문자메세지
# @app.route('/api/ajax_visit_emergency_approve_checkbox', methods=['POST'])
# def ajax_visit_emergency_approve_checkbox():
#     data = request.get_json()
#     data_length = len(data['checked_datas'])

#     if data_length < 1:
#         return "No Select"

#     for checked_data in data['checked_datas']:
#         visitor = Visitor.query.filter_by(id=checked_data).first()
#         visitor.approve = 1
#         visitor.exit = 0

#         today = date.today()
#         year = Year.query.filter_by(year=today.year).first()
#         if not year:
#             year = Year(year=today.year)
#             db.session.add(year)

#         month = Month.query.filter_by(year=today.year, month=today.month).first()
#         if not month:
#             month = Month(year=today.year, month=today.month)
#             db.session.add(month)
#             year.count += 1

#         day = Day.query.filter_by(year=today.year, month=today.month, day=today.day).first()
#         if not day:
#             day = Day(year=today.year, month=today.month, day=today.day, count=1)
#             db.session.add(day)
#             month.count += 1
#         else:
#             year.count += 1
#             month.count += 1
#             day.count += 1

#         db.session.commit()
#     return jsonify(result = "success")

# 방문객 관리 페이지 - 담당자 업데이트 체크 박스 api
@app.route('/api/ajax_index_manager_update_checkbox', methods=['POST'])
@login_required
def ajax_index_manager_update_checkbox():
    data = request.get_json()
    manager = data['manager']
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"
    if manager is None or manager == "":
        return "Error"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        visitor.manager = manager
        db.session.commit()
    return jsonify(result = "success")

# 방문객 관리 페이지 - 비고 업데이트 체크 박스 api
@app.route('/api/ajax_remarks_update_checkbox', methods=['POST'])
@login_required
def ajax_remarks_update_checkbox():
    data = request.get_json()
    remarks = data['remarks']
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"
    if remarks is None or remarks == "":
        return "Error"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        visitor.remarks = remarks
        db.session.commit()
    return jsonify(result = "success")

# 상황실 방문객 관리 페이지 - 세부 작업 위치 업데이트 체크 박스 api
@app.route('/api/ajax_detail_location_update_checkbox', methods=['POST'])
@login_required
def ajax_detail_location_update_checkbox():
    data = request.get_json()
    detail_location = data['detail_location']
    data_length = len(data['checked_datas'])

    if data_length < 1:
        return "No Select"
    if detail_location is None or detail_location == "":
        return "Error"

    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        visitor.detail_location = detail_location
        db.session.commit()
    return jsonify(result = "success")

# 내방객 관리 페이지 - 카드 불출 체크 박스 api
@app.route('/api/ajax_index_card_checkbox', methods=['POST'])
@login_required
def ajax_index_card_checkbox():
    data = request.get_json()
    card = data['card'].split(' ')
    print(card[0], card[1])
    data_length = len(data['checked_datas']) # 선택된 체크박스 수

    # 선택한 카드가 없거나 2개 이상 선택됐을 때 오류 발생
    if card is None:
        return "No Card"
    if data_length < 1:
        return "No Select"
    if data_length != 1:
        return "Multi Check"

    card_table = Card.query.filter_by(card_type=card[0], card_num=card[1]).first()
    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.exit == 1:
            return "Exited"
        if visitor.card_id != None:
            return "Use Card"

        visitor.card_id = card_table.id
        card_table.card_status = "불출"
        db.session.commit()
    return jsonify(result = "success")

# 내방객 관리 페이지 - 카드 회수 체크 박스 api
@app.route('/api/ajax_update_visit_recall_card_checkbox', methods=['POST'])
@login_required
def ajax_update_visit_recall_card_checkbox():
    data = request.get_json()
    checked_datas = data['checked_datas']
    data_length = len(data['checked_datas']) # 선택된 체크박스 수
    print(checked_datas)

    # 선택한 카드가 없음
    if data_length < 1:
        return "No Select"
    
    for checked_data in checked_datas:
        recall_visitor = Visitor.query.filter_by(id=checked_data).first()
        if recall_visitor.card_id == None:
            return "No Card"
        if recall_visitor.exit == 1:
            return "Exited"
        recall_visitor.card.card_status = "회수"
        recall_visitor.card_id = None
        db.session.commit()
    return jsonify(result = "success")

# 내방객 등록, 내방객 관리 페이지 - 방문객 수정 api
@app.route('/api/ajax_update_manage_visit', methods=['POST'])
@login_required
def ajax_update_manage_visit():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = request.get_json()
    visitor = data['update_btn']
    update_visitor = Visitor.query.filter_by(id=visitor).first()
    if update_visitor.card_id != None:
        return "Use Card"
    else:
        # 개인정보 조회 로그 남기기
        inquiry_log = Privacy_log("조회", current_user.id, request.remote_addr, current_timestamp, "내방객 조회", update_visitor.name)
        db.session.add(inquiry_log)
        db.session.commit()
        update_visitor_info = [update_visitor.id, update_visitor.name, aes.decrypt(update_visitor.department), update_visitor.object, aes.decrypt(update_visitor.phone), update_visitor.manager, update_visitor.device, update_visitor.remarks, update_visitor.location, update_visitor.work, update_visitor.company_type, update_visitor.company, update_visitor.work_content, update_visitor.personal_computer, update_visitor.model_name, update_visitor.serial_number, update_visitor.pc_reason, update_visitor.work_division, update_visitor.customer, update_visitor.device_division, update_visitor.device_count]
        
        return jsonify(response=update_visitor_info)


# 내방객 관리 페이지 - 방문객 삭제 api
@app.route('/api/ajax_delete_manage_visit', methods=['POST'])
@login_required
def ajax_delete_manage_visit():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = request.get_json()
    visitor = data['delete_btn']
    delete_visitor = Visitor.query.filter_by(id=visitor).first()
    if delete_visitor.card_id != None:
        return "Use Card"
    # 삭제 로그 남기기
    delete_log = Privacy_log("삭제", current_user.id, request.remote_addr, current_timestamp, "내방객 삭제", delete_visitor.name)
    db.session.add(delete_log)
    db.session.delete(delete_visitor)
    db.session.commit()
    return jsonify()


#===================================================================================


#===================================================================================

# MMS-IMAGE TEST SMS 문자 메세지 보내기 - 현장 등록 승인
def image_send_sms_current(name, date, object, location, manager, phone_num, device, work, company, work_content, personal_computer, model_name, work_division, device_division, device_count):
    connect_to_database()
    cursor = app.mysql_conn.cursor()  # 커서 생성
    insert_query = "INSERT INTO MMS_MSG (REQDATE, STATUS, TYPE, PHONE, CALLBACK, SUBJECT, MSG, FILE_CNT, FILE_TYPE1, FILE_PATH1) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

    if personal_computer == "반입" and work == "해당" and device == "반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-PC 반입: {personal_computer}\n"
            f"-모델명: {model_name}\n"
            f"-작업: {work}\n"
            f"-작업 분류: {work_division}\n"
            f"-작업위치: {location}\n"
            f"-요청 회사명: {company}\n"
            f"-작업내용: {work_content}\n"
            f"-장비반출입: {device}\n"
            f"-장비 기종: {device_division}\n"
            f"-장비 수량: {device_count}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "반입" and work == "해당" and device == "미반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-PC 반입: {personal_computer}\n"
            f"-모델명: {model_name}\n"
            f"-작업: {work}\n"
            f"-작업 분류: {work_division}\n"
            f"-작업위치: {location}\n"
            f"-요청 회사명: {company}\n"
            f"-작업내용: {work_content}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "반입" and work == "해당 없음" and device == "반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-PC 반입: {personal_computer}\n"
            f"-모델명: {model_name}\n"
            f"-장비반출입: {device}\n"
            f"-장비 기종: {device_division}\n"
            f"-장비 수량: {device_count}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "반입" and work == "해당 없음" and device == "미반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-PC 반입: {personal_computer}\n"
            f"-모델명: {model_name}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "미반입" and work == "해당" and device == "반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-작업: {work}\n"
            f"-작업 분류: {work_division}\n"
            f"-작업위치: {location}\n"
            f"-요청 회사명: {company}\n"
            f"-작업내용: {work_content}\n"
            f"-장비반출입: {device}\n"
            f"-장비 기종: {device_division}\n"
            f"-장비 수량: {device_count}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "미반입" and work == "해당" and device == "미반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-작업: {work}\n"
            f"-작업 분류: {work_division}\n"
            f"-작업위치: {location}\n"
            f"-요청 회사명: {company}\n"
            f"-작업내용: {work_content}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "미반입" and work == "해당 없음" and device == "반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-장비반출입: {device}\n"
            f"-장비 기종: {device_division}\n"
            f"-장비 수량: {device_count}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "미반입" and work == "해당 없음" and device == "미반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )

    qrcode_img = generate_qr_code(name, date)

    insert_data = (datetime.now(), '1', '0', phone_num, '0322110290', '[내방객 출입 관리 시스템 현장 등록 승인]', msg, '2', 'I', '/home/cjadmin/web/O_Pass/static/img/qrcode.jpg')  # 삽입할 데이터를 튜플로 정의
    cursor.execute(insert_query, insert_data)  # 쿼리 실행 및 데이터 전달
    app.mysql_conn.commit()  # 변경 사항 커밋
    cursor.close()  # 커서 닫기

# MMS-IMAGE TEST SMS 문자 메세지 보내기 - 사전 등록 승인
def image_send_sms_previous(name, date, object, location, manager, phone_num, device, work, company, work_content, personal_computer, model_name, work_division, device_division, device_count):
    connect_to_database()
    cursor = app.mysql_conn.cursor()  # 커서 생성
    insert_query = "INSERT INTO MMS_MSG (REQDATE, STATUS, TYPE, PHONE, CALLBACK, SUBJECT, MSG, FILE_CNT, FILE_TYPE1, FILE_PATH1) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

    if personal_computer == "반입" and work == "해당" and device == "반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-PC 반입: {personal_computer}\n"
            f"-모델명: {model_name}\n"
            f"-작업: {work}\n"
            f"-작업 분류: {work_division}\n"
            f"-작업위치: {location}\n"
            f"-요청 회사명: {company}\n"
            f"-작업내용: {work_content}\n"
            f"-장비반출입: {device}\n"
            f"-장비 기종: {device_division}\n"
            f"-장비 수량: {device_count}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "반입" and work == "해당" and device == "미반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-PC 반입: {personal_computer}\n"
            f"-모델명: {model_name}\n"
            f"-작업: {work}\n"
            f"-작업 분류: {work_division}\n"
            f"-작업위치: {location}\n"
            f"-요청 회사명: {company}\n"
            f"-작업내용: {work_content}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "반입" and work == "해당 없음" and device == "반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-PC 반입: {personal_computer}\n"
            f"-모델명: {model_name}\n"
            f"-장비반출입: {device}\n"
            f"-장비 기종: {device_division}\n"
            f"-장비 수량: {device_count}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "반입" and work == "해당 없음" and device == "미반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-PC 반입: {personal_computer}\n"
            f"-모델명: {model_name}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "미반입" and work == "해당" and device == "반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-작업: {work}\n"
            f"-작업 분류: {work_division}\n"
            f"-작업위치: {location}\n"
            f"-요청 회사명: {company}\n"
            f"-작업내용: {work_content}\n"
            f"-장비반출입: {device}\n"
            f"-장비 기종: {device_division}\n"
            f"-장비 수량: {device_count}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "미반입" and work == "해당" and device == "미반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-작업: {work}\n"
            f"-작업 분류: {work_division}\n"
            f"-작업위치: {location}\n"
            f"-요청 회사명: {company}\n"
            f"-작업내용: {work_content}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "미반입" and work == "해당 없음" and device == "반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-장비반출입: {device}\n"
            f"-장비 기종: {device_division}\n"
            f"-장비 수량: {device_count}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    elif personal_computer == "미반입" and work == "해당 없음" and device == "미반입":
        msg = (
            f"{name}님 안녕하세요.\n"
            f"송도 IDC 센터에 방문하신 것을 환영합니다.\n"
            f"-등록시간: {date}\n"
            f"-방문위치: 인천광역시 연수구 하모니로177번길 20\n"
            f"-방문목적: {object}\n"
            f"-담당자: {manager}\n"
            f"-QR Code◀"
        )
    
    generate_qr_code(name, date)

    insert_data = (datetime.now(), '1', '0', phone_num, '0322110290', '[내방객 출입 관리 시스템 사전 등록 승인]', msg, '2', 'I', '/home/cjadmin/web/O_Pass/static/img/qrcode.jpg')  # 삽입할 데이터를 튜플로 정의
    cursor.execute(insert_query, insert_data)  # 쿼리 실행 및 데이터 전달
    app.mysql_conn.commit()  # 변경 사항 커밋
    cursor.close()  # 커서 닫기

# manage qr 재전송 api
@app.route('/api/ajax_manage_qrcode_send', methods=['POST'])
@login_required
def ajax_manage_qrcode_send():
    data = request.get_json()
    qrcode_id = data['qrcode_btn']
    print(qrcode_id)
    qrcode_visitor = Visitor.query.filter_by(id=qrcode_id).first()
    print(qrcode_visitor.phone)
    if qrcode_visitor.card_id:
        return "Use Card"
    
    if qrcode_visitor.personal_computer == 0:
        personal_computer = "미반입"
    else:
        personal_computer = "반입"

    if qrcode_visitor.device == 0:
        device = "미반입"
    else:
        device = "반입"

    if qrcode_visitor.work == 0:
        work = "해당 없음"
    else:
        work = "해당"
    
    if qrcode_visitor.registry == "사전 등록":
        image_send_sms_previous(qrcode_visitor.name, qrcode_visitor.approve_date, qrcode_visitor.object, qrcode_visitor.location, qrcode_visitor.manager, aes.decrypt(qrcode_visitor.phone), device, work, qrcode_visitor.company, qrcode_visitor.work_content, personal_computer, qrcode_visitor.model_name, qrcode_visitor.work_division, qrcode_visitor.device_division, qrcode_visitor.device_count)
    else:
        image_send_sms_current(qrcode_visitor.name, qrcode_visitor.approve_date, qrcode_visitor.object, qrcode_visitor.location, qrcode_visitor.manager, aes.decrypt(qrcode_visitor.phone), device, work, qrcode_visitor.company, qrcode_visitor.work_content, personal_computer, qrcode_visitor.model_name, qrcode_visitor.work_division, qrcode_visitor.device_division, qrcode_visitor.device_count)
    return jsonify()

#===================================================================================


#===================================================================================

# # Excel 다운로드 api
@app.route('/api/ajax_excel_download_1', methods=['GET', 'POST'])
@login_required
def ajax_excel_download_1():
    data = request.get_json()
    select_option = data['excel_option_val']
    option_text = data['option_text_val']
    start_date = data['start_date_val']
    end_date = data['end_date_val']

    print(select_option, option_text, start_date, end_date)
    if option_text == '':
        return "No Input"
    if start_date == '' or end_date == '':
        return "No Date"
    if start_date > end_date:
        return "Date Error"

    # 내방객 점검 일지
    file_name = 'excel.xlsx'
    
    # 파일 읽기
    workbook = openpyxl.load_workbook(file_name)  # 기존 파일을 로드합니다.
    sheet = workbook.active
    # 데이터 쓰기
    if select_option == 'name':
        if start_date == end_date:
            # If start_date and end_date are the same, include only a single day
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            start_datetime_str = start_datetime.strftime('%Y-%m-%d')
            exited_visitors = Visitor.query.filter(
                and_(
                    Visitor.exit_date >= start_datetime_str,
                    Visitor.exit_date < (start_datetime + timedelta(days=1)).strftime('%Y-%m-%d'),
                    Visitor.name == option_text
                )
            )
        else:
            # Include the range between start_date and end_date
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            start_datetime_str = start_datetime.strftime('%Y-%m-%d')
            end_datetime_str = end_datetime.strftime('%Y-%m-%d')
            if start_datetime != end_datetime:
                end_datetime += timedelta(days=1)  # Add one day to include the end_date
            exited_visitors = Visitor.query.filter(
                and_(
                    Visitor.exit_date >= start_datetime_str,
                    Visitor.exit_date <= end_datetime.strftime('%Y-%m-%d'),
                    Visitor.name == option_text
                )
            )
    elif select_option == 'department':
        if start_date == end_date:
            # If start_date and end_date are the same, include only a single day
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            start_datetime_str = start_datetime.strftime('%Y-%m-%d')
            exited_visitors = Visitor.query.filter(
                and_(
                    Visitor.exit_date >= start_datetime_str,
                    Visitor.exit_date < (start_datetime + timedelta(days=1)).strftime('%Y-%m-%d'),
                    Visitor.department == aes.encrypt(option_text)
                )
            )
        else:
            # Include the range between start_date and end_date
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            start_datetime_str = start_datetime.strftime('%Y-%m-%d')
            end_datetime_str = end_datetime.strftime('%Y-%m-%d')
            if start_datetime != end_datetime:
                end_datetime += timedelta(days=1)  # Add one day to include the end_date
            exited_visitors = Visitor.query.filter(
                and_(
                    Visitor.exit_date >= start_datetime_str,
                    Visitor.exit_date <= end_datetime.strftime('%Y-%m-%d'),
                    Visitor.department == aes.encrypt(option_text)
                )
            )
    elif select_option == 'phone':
        if start_date == end_date:
            # If start_date and end_date are the same, include only a single day
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            start_datetime_str = start_datetime.strftime('%Y-%m-%d')
            exited_visitors = Visitor.query.filter(
                and_(
                    Visitor.exit_date >= start_datetime_str,
                    Visitor.exit_date < (start_datetime + timedelta(days=1)).strftime('%Y-%m-%d'),
                    Visitor.phone == aes.encrypt(option_text)
                )
            )
        else:
            # Include the range between start_date and end_date
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            start_datetime_str = start_datetime.strftime('%Y-%m-%d')
            end_datetime_str = end_datetime.strftime('%Y-%m-%d')
            if start_datetime != end_datetime:
                end_datetime += timedelta(days=1)  # Add one day to include the end_date
            exited_visitors = Visitor.query.filter(
                and_(
                    Visitor.exit_date >= start_datetime_str,
                    Visitor.exit_date <= end_datetime.strftime('%Y-%m-%d'),
                    Visitor.phone == aes.encrypt(option_text)
                )
            )
    elif select_option == 'manager':
        if start_date == end_date:
            # If start_date and end_date are the same, include only a single day
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            start_datetime_str = start_datetime.strftime('%Y-%m-%d')
            exited_visitors = Visitor.query.filter(
                and_(
                    Visitor.exit_date >= start_datetime_str,
                    Visitor.exit_date < (start_datetime + timedelta(days=1)).strftime('%Y-%m-%d'),
                    Visitor.manager == option_text
                )
            )
        else:
            # Include the range between start_date and end_date
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            start_datetime_str = start_datetime.strftime('%Y-%m-%d')
            end_datetime_str = end_datetime.strftime('%Y-%m-%d')
            if start_datetime != end_datetime:
                end_datetime += timedelta(days=1)  # Add one day to include the end_date
            exited_visitors = Visitor.query.filter(
                and_(
                    Visitor.exit_date >= start_datetime_str,
                    Visitor.exit_date <= end_datetime.strftime('%Y-%m-%d'),
                    Visitor.manager == option_text
                )
            )
    else:
        return "No Data"

    # 열 너비 설정 (A, B, C 순서로 30)
    column_widths = [5, 10, 20, 20, 15, 15, 15, 15, 15, 10, 15, 20, 20, 10, 15, 20, 15, 18, 15, 10, 15, 15, 15, 15]  # 각 열의 너비를 30으로 설정

    for col_num, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = width

    # 헤더 생성
    header = ["No", "등록", "출입 시간", "퇴실 시간", "이름", "소속", "전화번호", "담당자", "방문목적",
            "PC 반입", "모델명", "시리얼번호", "반입사유", "작업", "작업 분류", "작업 내용", "작업 위치",
            "작업 요청 회사 타입", "작업 요청 회사명", "장비 반입", "고객사", "장비 분류", "장비 수량", "비고"]
    
    # 헤더 추가
    sheet.append(header)

    # A1 셀 가운데 정렬
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    for row_num, exited_visitor in enumerate(exited_visitors, 2):
        # 이름에서 두 번째 글자를 '*'로 마스킹 처리
        name = exited_visitor.name
        phone = aes.decrypt(exited_visitor.phone)
        department = aes.decrypt(exited_visitor.department)
        personal_computer = exited_visitor.personal_computer
        work = exited_visitor.work
        device = exited_visitor.device
        if len(name) >= 2:
            masked_name = name[0] + '*' + name[2:]
        else:
            masked_name = name
        if len(phone) >= 5:
            masked_phone = phone[:3] + '****' + phone[7:]
        else:
            masked_phone = phone
        if personal_computer == 1:
            personal_computer = 'O'
        else:
            personal_computer = 'X'
        if work == 1:
            work = 'O'
        else:
            work = 'X'
        if device == 1:
            device = 'O'
        else:
            device = 'X'
        
        row = (
            row_num-1, exited_visitor.registry, exited_visitor.approve_date, exited_visitor.exit_date, masked_name, department, masked_phone, 
            exited_visitor.manager, exited_visitor.object, personal_computer, exited_visitor.model_name, exited_visitor.serial_number, exited_visitor.pc_reason,
            work, exited_visitor.work_division, exited_visitor.work_content, exited_visitor.location, exited_visitor.company_type, exited_visitor.company,
            device, exited_visitor.customer, exited_visitor.device_division, exited_visitor.device_count, exited_visitor.remarks,
        )
        sheet.append(row)

        # 헤더를 A1부터 순서대로 열 갯수만큼 가운데 정렬하면서 추가
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[col_letter + '1']
            cell.value = header_text
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 헤더 스타일 설정
            cell.font = Font(color='FFFFFFFF', bold=True)  # 글자색을 하얀색으로, 볼드체로 설정
            cell.fill = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid')  # 배경색을 진한 회색으로 설정

        for col_num, _ in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[col_letter + str(row_num)]
            cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save('excel/' + start_date + '-' + end_date + ' 내방객 출입점검 일지1.xlsx')
    return jsonify(result="success")

@app.route('/api/ajax_excel_download_2', methods=['GET', 'POST'])
@login_required
def ajax_excel_download_2():
    data = request.get_json()
    start_date = data['start_date_val']
    end_date = data['end_date_val']

    print(start_date, end_date)
    if start_date == '' or end_date == '':
        return "No Date"
    if start_date > end_date:
        return "Date Error"

    # 내방객 점검 일지
    file_name = 'excel.xlsx'
    
    # 파일 읽기
    workbook = openpyxl.load_workbook(file_name)  # 기존 파일을 로드합니다.
    sheet = workbook.active
    # 데이터 쓰기
    if start_date == end_date:
        # If start_date and end_date are the same, include only a single day
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        start_datetime_str = start_datetime.strftime('%Y-%m-%d')
        exited_visitors = Visitor.query.filter(
            and_(
                Visitor.exit_date >= start_datetime_str,
                Visitor.exit_date < (start_datetime + timedelta(days=1)).strftime('%Y-%m-%d')
            )
        )
    else:
        # Include the range between start_date and end_date
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
        start_datetime_str = start_datetime.strftime('%Y-%m-%d')
        end_datetime_str = end_datetime.strftime('%Y-%m-%d')
        if start_datetime != end_datetime:
            end_datetime += timedelta(days=1)  # Add one day to include the end_date
        exited_visitors = Visitor.query.filter(
            and_(
                Visitor.exit_date >= start_datetime_str,
                Visitor.exit_date <= end_datetime.strftime('%Y-%m-%d')
            )
        )

    # 열 너비 설정 (A, B, C 순서로 30)
    column_widths = [5, 10, 20, 20, 15, 15, 15, 15, 15, 10, 15, 20, 20, 10, 15, 20, 15, 18, 15, 10, 15, 15, 15, 15]  # 각 열의 너비를 30으로 설정

    for col_num, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = width

    # 헤더 생성
    header = ["No", "등록", "출입 시간", "퇴실 시간", "이름", "소속", "전화번호", "담당자", "방문목적",
            "PC 반입", "모델명", "시리얼번호", "반입사유", "작업", "작업 분류", "작업 내용", "작업 위치",
            "작업 요청 회사 타입", "작업 요청 회사명", "장비 반입", "고객사", "장비 분류", "장비 수량", "비고"]
    
    # 헤더 추가
    sheet.append(header)

    # A1 셀 가운데 정렬
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    for row_num, exited_visitor in enumerate(exited_visitors, 2):
        # 이름에서 두 번째 글자를 '*'로 마스킹 처리
        name = exited_visitor.name
        phone = aes.decrypt(exited_visitor.phone)
        department = aes.decrypt(exited_visitor.department)
        personal_computer = exited_visitor.personal_computer
        work = exited_visitor.work
        device = exited_visitor.device
        if len(name) >= 2:
            masked_name = name[0] + '*' + name[2:]
        else:
            masked_name = name
        if len(phone) >= 5:
            masked_phone = phone[:3] + '****' + phone[7:]
        else:
            masked_phone = phone
        if personal_computer == 1:
            personal_computer = 'O'
        else:
            personal_computer = 'X'
        if work == 1:
            work = 'O'
        else:
            work = 'X'
        if device == 1:
            device = 'O'
        else:
            device = 'X'
        
        row = (
            row_num-1, exited_visitor.registry, exited_visitor.approve_date, exited_visitor.exit_date, masked_name, department, masked_phone, 
            exited_visitor.manager, exited_visitor.object, personal_computer, exited_visitor.model_name, exited_visitor.serial_number, exited_visitor.pc_reason,
            work, exited_visitor.work_division, exited_visitor.work_content, exited_visitor.location, exited_visitor.company_type, exited_visitor.company,
            device, exited_visitor.customer, exited_visitor.device_division, exited_visitor.device_count, exited_visitor.remarks,
        )
        sheet.append(row)

        # 헤더를 A1부터 순서대로 열 갯수만큼 가운데 정렬하면서 추가
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[col_letter + '1']
            cell.value = header_text
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 헤더 스타일 설정
            cell.font = Font(color='FFFFFFFF', bold=True)  # 글자색을 하얀색으로, 볼드체로 설정
            cell.fill = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid')  # 배경색을 진한 회색으로 설정

        for col_num, _ in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[col_letter + str(row_num)]
            cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save('excel/' + start_date + '-' + end_date + ' 내방객 출입점검 일지1.xlsx')
    return jsonify(result="success")

#===================================================================================


#===================================================================================

# 카드 종류 추가 api
@app.route('/api/ajax_card_type_create', methods=['POST'])
@login_required
def ajax_card_type_create():
    data = request.get_json()
    print(data['card_type_val'])
    card_type_value = data['card_type_val']
    
    card_type_distinct = Card.query.filter_by(card_type=card_type_value).all()
    if card_type_distinct:
        return "type distinct"

    if card_type_value:
        card_type = Card(card_type_value, 1, '회수')
        db.session.add(card_type)
        db.session.commit()
    
    return jsonify(result="success")

# cards 카드 수량 추가 api
@app.route('/api/ajax_add_card', methods=['POST'])
@login_required
def ajax_add_card():
    data = request.get_json()
    card_number = data['add_card_value']
    card_type = data['select_card_type']
    print(card_type, card_number)

    if card_number == 0:
        return "No Number"
    
    cards = Card.query.filter_by(card_type=card_type).order_by(func.cast(Card.card_num, Integer).desc()).first()
    for index in range(1, int(card_number) + 1):
        add_cards = Card(card_type, int(cards.card_num)+int(index), '회수')
        db.session.add(add_cards)
        db.session.commit()
        
    print(cards.card_num)
    print(cards)
    return jsonify(result = "success")

# cards 카드 선택 추가 api
@app.route('/api/ajax_add_card_select', methods=['POST'])
@login_required
def ajax_add_card_select():
    data = request.get_json()
    card_number = data['add_card_value']
    card_type = data['select_card_type']
    print(card_type, card_number)
    
    card = Card.query.filter_by(card_type=card_type, card_num=card_number).first()
    if card != None:
        return "Use Card"
    
    add_card = Card(card_type, card_number, '회수')
    db.session.add(add_card)
    db.session.commit()

    return jsonify(result = "success")

# cards 카드 제거 api
@app.route('/api/ajax_card_delete', methods=['POST'])
@login_required
def ajax_card_delete():
    data = request.get_json()
    delete_id = data['delete_id']
    print(delete_id)

    delete_card = Card.query.filter_by(id=delete_id).first()
    db.session.delete(delete_card)
    db.session.commit()
    return jsonify(result = "success")

# 카드 분실 api
@app.route('/api/ajax_card_lose', methods=['POST'])
@login_required
def ajax_card_lose():
    data = request.get_json()
    card = data['card'].split(' ')
    card_lose = Card.query.filter_by(card_type=card[0], card_num=card[1]).first()
    card_lose.card_status = '분실'
    db.session.commit()
    
    return jsonify(result = "success")

# 카드 초기화 api
@app.route('/api/ajax_card_reset', methods=['POST'])
@login_required
def ajax_card_reset():
    # Card 테이블 초기화
    db.session.query(Card).delete()
    db.session.commit()
    return jsonify(result = "success")

#===================================================================================


#===================================================================================
# 키 종류 추가 api
@app.route('/api/ajax_key_type_create', methods=['POST'])
@login_required
def ajax_key_type_create():
    data = request.get_json()
    key_type_value = data['key_type_val']
    
    key_type_distinct = Rack.query.filter_by(key_type=key_type_value).all()
    if key_type_distinct:
        return "type distinct"

    if key_type_value:
        key_type = Rack(key_type_value, 1, '회수')
        db.session.add(key_type)
        db.session.commit()
    
    return jsonify(result="success")

# keys 키 수량 추가 api
@app.route('/api/ajax_add_key', methods=['POST'])
@login_required
def ajax_add_key():
    data = request.get_json()
    key_number = data['add_key_value']
    key_type = data['select_key_type']

    if key_number == 0:
        return "No Number"
    
    keys = Rack.query.filter_by(key_type=key_type).order_by(func.cast(Rack.key_num, Integer).desc()).first()
    for index in range(1, int(key_number) + 1):
        add_keys = Rack(key_type, int(keys.key_num)+int(index), '회수')
        db.session.add(add_keys)
        db.session.commit()
        
    return jsonify(result = "success")

# keys 키 선택 추가 api
@app.route('/api/ajax_add_key_select', methods=['POST'])
@login_required
def ajax_add_key_select():
    data = request.get_json()
    key_number = data['add_key_value']
    key_type = data['select_key_type']
    
    key = Rack.query.filter_by(key_type=key_type, key_num=key_number).first()
    if key != None:
        return "Use Key"
    
    add_key = Rack(key_type, key_number, '회수')
    db.session.add(add_key)
    db.session.commit()

    return jsonify(result = "success")

# 키 분실 api
@app.route('/api/ajax_key_lose', methods=['POST'])
@login_required
def ajax_key_lose():
    data = request.get_json()
    key = data['key'].split(' ')
    key_lose = Rack.query.filter_by(key_type=key[0], key_num=key[1]).first()
    key_lose.key_status = '분실'
    db.session.commit()
    
    return jsonify(result = "success")

# keys 키 제거 api
@app.route('/api/ajax_key_delete', methods=['POST'])
@login_required
def ajax_key_delete():
    data = request.get_json()
    delete_id = data['delete_id']

    delete_key = Rack.query.filter_by(id=delete_id).first()
    db.session.delete(delete_key)
    db.session.commit()
    return jsonify(result = "success")

# 키 초기화 api
@app.route('/api/ajax_key_reset', methods=['POST'])
@login_required
def ajax_key_reset():
    # 키 테이블 초기화
    db.session.query(Rack).delete()
    db.session.commit()
    return jsonify(result = "success")

#===================================================================================

# 유저 프로필 페이지
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    if request.method == 'POST':
        department = request.form['current_department']
        rank = request.form['current_rank']

        user = User.query.filter_by(email=current_user.email).first()
        if department and rank:
            user.department = department
            user.rank = rank
            db.session.commit()
            flash("프로필이 업데이트 되었습니다.")
            return redirect('profile')
        else:
            flash("프로필 업데이트에 실패했습니다.")
            return redirect('profile')
    return render_template('user_profile.html')

# 유저 인증 페이지
@app.route('/authenticated', methods=['GET', 'POST'])
@login_required
def user_authenticated():
    if request.method == 'POST':
        password = request.form['current_pwd']

        user = User.query.filter_by(email=current_user.email).first()
        if user:
            if bcrypt.check_password_hash(user.password, password):
                result = hashlib.sha256(user.username.encode())
                return redirect(url_for('user_profile_update', _method='POST'))
            else:
                flash("비밀번호를 확인 후 다시 입력해주세요.")
                return redirect('authenticated')
        else:
            flash("다시 입력해주세요.")
            return redirect('authenticated')
    else:
        return render_template('authenticated.html')


# 유저 정보 변경 페이지
@app.route('/profile_update', methods=['GET', 'POST'])
@login_required
def user_profile_update():
    korean_timezone = pytz.timezone('Asia/Seoul')
    if request.method == 'POST':
        current_password = request.form['current_pwd']
        new_password_1 = request.form['new_pwd_1']
        new_password_2 = request.form['new_pwd_2']
        
        user = User.query.filter_by(email=current_user.email).first()

        current_time_korean = datetime.now(korean_timezone).strftime("%Y-%m-%d %H:%M:%S")
        fomatting_time_korean = datetime.strptime(current_time_korean, '%Y-%m-%d %H:%M:%S')
        time_since_registration = fomatting_time_korean - user.registered_at

        if user:
            if user.attempts == 'attempts_password':
                if bcrypt.check_password_hash(user.password, current_password):
                    if current_password == new_password_1:
                        flash("현재 비밀번호와 새 비밀번호는 같을 수 없습니다.")
                    elif new_password_1 != new_password_2:
                        flash("비밀번호와 비밀번호재입력이 서로 다릅니다.")
                    elif len(new_password_1) < 8:
                        flash("비밀번호는 8자 이상이어야 합니다.")
                    elif len(new_password_1) > 14:
                        flash("비밀번호는 14자 이하여야 합니다.")
                    else:
                        # 최소 3종류 이상 포함하는지 검사
                        categories = 0
                        if re.search(r'[A-Z]', new_password_1):
                            categories += 1
                        if re.search(r'[a-z]', new_password_1):
                            categories += 1
                        if re.search(r'\d', new_password_1):
                            categories += 1
                        if re.search(r'[!@#$%^&*(),.?":{}|<>]', new_password_1):
                            categories += 1
                        if categories < 4:
                            flash('비밀번호 3종 복잡도를 만족하지 않습니다.')
                        elif re.search(r'(.)\1{3,}', new_password_1.lower()) or contains_consecutive(new_password_1, 4) or contains_decreasing(new_password_1, 4) or is_keyboard_consecutive(new_password_1.lower(), 4):
                            flash('4자 이상의 연속 문자를 사용할 수 없습니다.')
                        else:
                            # 비밀번호 수정
                            hashed_password = bcrypt.generate_password_hash(new_password_1)
                            
                            # 비밀번호 이력 확인
                            confirm_password = Password_log.query.filter_by(user_id=user.id).order_by(Password_log.id.desc()).limit(3).all()
                            for idx in confirm_password:
                                if bcrypt.check_password_hash(idx.password_log, new_password_1):
                                    flash('최근 3개의 비밀번호는 사용하실 수 없습니다.')
                                    return redirect('profile_update')

                            user.password = hashed_password
                            user.password_changed_at = current_time_korean
                            user.attempts = None

                            # 비밀번호 이력 보관
                            password_log = Password_log(hashed_password, user.id)
                            password_change_log = Password_change_log(user.email, fomatting_time_korean, request.remote_addr, user.id)
                            db.session.add(password_log)
                            db.session.add(password_change_log)
                            db.session.commit()

                            logout_user()
                            flash("비밀번호가 정상적으로 변경되었습니다.")
                            return redirect('login')
                    
            elif time_since_registration < timedelta(days=1):
                flash("회원가입 후 최소 1일이 지나야 비밀번호를 수정할 수 있습니다.")

            else:
                if bcrypt.check_password_hash(user.password, current_password):
                    if current_password == new_password_1:
                        flash("현재 비밀번호와 새 비밀번호는 같을 수 없습니다.")
                    elif new_password_1 != new_password_2:
                        flash("비밀번호와 비밀번호재입력이 서로 다릅니다.")
                    elif len(new_password_1) < 8:
                        flash("비밀번호는 8자 이상이어야 합니다.")
                    elif len(new_password_1) > 14:
                        flash("비밀번호는 14자 이하여야 합니다.")
                    else:
                        # 최소 3종류 이상 포함하는지 검사
                        categories = 0
                        if re.search(r'[A-Z]', new_password_1):
                            categories += 1
                        if re.search(r'[a-z]', new_password_1):
                            categories += 1
                        if re.search(r'\d', new_password_1):
                            categories += 1
                        if re.search(r'[!@#$%^&*(),.?":{}|<>]', new_password_1):
                            categories += 1
                        if categories < 4:
                            flash('비밀번호 3종 복잡도를 만족하지 않습니다.')
                        elif re.search(r'(.)\1{3,}', new_password_1.lower()) or contains_consecutive(new_password_1, 4) or contains_decreasing(new_password_1, 4) or is_keyboard_consecutive(new_password_1.lower(), 4):
                            flash('4자 이상의 연속 문자를 사용할 수 없습니다.')
                        else:
                            # 비밀번호 수정
                            hashed_password = bcrypt.generate_password_hash(new_password_1)
                            print(hashed_password)

                            # 비밀번호 이력 확인
                            confirm_password = Password_log.query.filter_by(user_id=user.id).order_by(Password_log.id.desc()).limit(3).all()
                            for idx in confirm_password:
                                if bcrypt.check_password_hash(idx.password_log, new_password_1):
                                    flash('최근 3개의 비밀번호는 사용하실 수 없습니다.')
                                    return redirect('profile_update')
                                
                            user.password = hashed_password
                            user.password_changed_at = current_time_korean

                            # 비밀번호 이력 보관
                            password_log = Password_log(hashed_password, user.id)
                            password_change_log = Password_change_log(user.email, fomatting_time_korean, request.remote_addr, user.id)
                            db.session.add(password_log)
                            db.session.add(password_change_log)
                            db.session.commit()

                            logout_user()
                            flash("비밀번호가 정상적으로 변경되었습니다.")
                            return redirect('login')
                else:
                    flash("비밀번호를 잘못 입력하셨습니다.")

    return render_template('user_profile_update.html')

#===================================================================================

# 방문객 현장 등록 페이지
@app.route('/form-input', methods=['GET','POST'])
def form_input():
    if request.method == 'POST':
        date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # 필수 입력
        name = request.form['visitorName']
        department = request.form['visitorDepartment']
        phone = request.form['visitorPhone']
        manager = request.form['visitorManager']
        object = request.form['visitorObject']
        personal_computer = request.form['visitorPC']
        work = request.form['visitorWork']
        device = request.form['visitorDevice']
        

        # 선택 입력
        model_name = request.form.get('visitorModelName')
        serial_number = request.form.get('visitorSerialNumber')
        pc_reason = request.form.get('visitorReason')

        work_division = request.form.get('visitorWorkDivision')
        work_content = request.form.get('visitorWorkContent')
        location = request.form.get('visitorLocation')
        company = request.form.get('visitorCompanyName')

        customer = request.form.get('visitorCustomer')
        device_division = request.form.get('visitorDeviceDivision')
        device_count = request.form.get('visitorDeviceCount')
        remarks = request.form.get('visitorRemarks')

        if personal_computer == '반입':
            personal_computer = True
        else:
            personal_computer = False

        if work == '해당':
            work = True
        else:
            work = False

        if device == '반입':
            device = True
        else:
            device = False

        privacy = Privacy(name, aes.encrypt(department), aes.encrypt(phone), manager, device, work, remarks, object, location, "", company, work_content, date, "현장 등록", personal_computer, model_name, serial_number, pc_reason, work_division, customer, device_division, device_count)
        visitor = Visitor(name, aes.encrypt(department), aes.encrypt(phone), location, manager, device, remarks, object, date, 0, "현장 등록", work, "", company, work_content, 0, personal_computer, model_name, serial_number, pc_reason, work_division, customer, device_division, device_count)
        db.session.add(privacy)
        db.session.add(visitor)
        db.session.commit()
        return redirect('form')
    return render_template('form-input.html')
# def __init__(self, name, department, phone, location, manager, device, remarks, object, created_time, approve, registry, work, company_type, company, work_content, detail_location):

#===================================================================================

# 키 불출 체크 박스 api
@app.route('/api/ajax_use_rack_key_checkbox', methods=['POST'])
@login_required
def ajax_use_rack_key_checkbox():
    data = request.get_json()
    key = data['key'].split(' ')
    print(key[0], key[1])
    data_length = len(data['checked_datas']) # 선택된 체크박스 수

    # 선택한 키가 없을 때 선택하지 않았을 때 오류 발생
    if key is None:
        return "No Key"
    if data_length < 1:
        return "No Select"

    key_table = Rack.query.filter_by(key_type=key[0], key_num=key[1]).first()
    for checked_data in data['checked_datas']:
        visitor = Visitor.query.filter_by(id=checked_data).first()
        if visitor.card_id == None:
            return "No Card"
        if visitor.exit == 1:
            return "Exited"
        if visitor.rack_id != None:
            return "Use Key"
        if visitor.work == False:
            return "No Work"

        visitor.rack_id = key_table.id
        key_table.key_status = "불출"
        db.session.commit()
    return jsonify(result = "success")

# 키 회수 체크 박스 api
@app.route('/api/ajax_recall_rack_key_checkbox', methods=['POST'])
@login_required
def ajax_recall_rack_key_checkbox():
    data = request.get_json()
    checked_datas = data['checked_datas']
    data_length = len(data['checked_datas']) # 선택된 체크박스 수

    # 선택한 키가 없음
    if data_length < 1:
        return "No Select"
    
    for checked_data in checked_datas:
        recall_visitor = Visitor.query.filter_by(id=checked_data).first()
        if recall_visitor.rack_id == None:
            return "No Key"
        if recall_visitor.exit == 1:
            return "Exited"
        recall_visitor.rack.key_status = "회수"
        recall_visitor.rack_id = None
        db.session.commit()
    return jsonify(result = "success")

#===================================================================================

@app.route('/api/user_delete', methods=['POST'])
@login_required
def user_delete():
    data = request.get_json()
    user_delete = User.query.filter_by(id=current_user.id).first()
    if user_delete:
        db.session.delete(user_delete)
        db.session.commit()
        return jsonify(result = "success")
    else:
        return "Error"

#===================================================================================
def generate_qr_code(name, date):
    try:
        qr_data = "https://opass.cj.net/qr_auth?id=" + name + "&date=" + date
        qr_img = qrcode.make(qr_data)
        save_path = 'static/img/qrcode.jpg'

        # 파일이 이미 존재하는 경우 덮어쓰기
        if os.path.exists(save_path):
            os.remove(save_path)

        qr_img.save(save_path)
    except Exception as e:
        print(e)

    return "QR Code generated"

@app.route('/qr_auth')
def authenticate():
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    name = request.args.get('id')
    date = request.args.get('date')
    visitor = Visitor.query.filter_by(name=name, approve_date=date).order_by(Visitor.id.desc()).first()

    if visitor:
        visitor.entry_date = current_timestamp
        db.session.commit()
        visitor_info = [visitor.registry, visitor.name, aes.decrypt(visitor.department), visitor.manager, visitor.created_date, visitor.approve_date, visitor.personal_computer, visitor.model_name, visitor.serial_number, visitor.pc_reason, visitor.work, visitor.work_division, visitor.work_content, visitor.location, visitor.company_type, visitor.company, visitor.device, visitor.customer, visitor.device_division, visitor.device_count, visitor.remarks, visitor.entry_date]
        return render_template('user_authenticated.html', visitor=visitor_info)
    else:
        flash('승인된 내방객이 아닙니다.')
        return "인증 실패"


@app.errorhandler(jinja2.exceptions.TemplateNotFound)
def template_not_found(e):
    return not_found(e)

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html')

# MySQL 데이터베이스에 연결하는 함수
def connect_to_database():
    db_user, db_password, db_host, db_port, db_name = db_connector()
    app.config['DB_HOST'] = db_host
    app.config['DB_USER'] = db_user
    app.config['DB_PASSWORD'] = db_password
    app.config['DB_DATABASE'] = db_name
    app.config['DB_PORT'] = db_port
    app.mysql_conn = mysql.connector.connect(
        host=app.config['DB_HOST'],
        user=app.config['DB_USER'],
        password=app.config['DB_PASSWORD'],
        database=app.config['DB_DATABASE'],
        port=app.config['DB_PORT']
    )

if __name__ == '__main__':
    connect_to_database()
    # csrf.init_app(app) # CSRF Config
    app.run(debug=True)
